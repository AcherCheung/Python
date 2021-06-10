from random import choice
from time import ctime, sleep, time
import openpyxl
from openpyxl import Workbook, load_workbook

# 设置文件地址
from openpyxl.utils import get_column_letter

addr = 'a/openpyxl.xlsx'
# 加载已存在的工作薄
wb = load_workbook(addr)
# 新建工作表
ws1 = wb.create_sheet('case_two')
# 第一行写入数据
ws1.append(['Time', 'Tittle', 'letter'])

# 输入内容
for i in range(1, 10):
    Time = ctime()
    Tittle = f'现在是 {time()}'
    letter = get_column_letter(choice(range(1, 50)))
    ws1.append([Time, Tittle, letter])
    sleep(0.5)

# 把上面写入的内容打印在控制台
for i in ws1.rows: #获取每一行数据
    for n in i:
        print(n.value, end='\t')
    print()
wb.save(addr)



